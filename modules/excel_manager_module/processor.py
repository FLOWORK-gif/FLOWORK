#######################################################################
# dev : awenk audico
# EMAIL SAHIDINAOLA@GMAIL.COM
# WEBSITE WWW.TEETAH.ART
# File NAME : C:\FLOWORK\modules\excel_manager_module\processor.py
# JUMLAH BARIS : 150
#######################################################################

import os
import pandas as pd
from flowork_kernel.api_contract import BaseModule, IExecutable, IConfigurableUI, IDataPreviewer
from flowork_kernel.exceptions import PermissionDeniedError
from flowork_kernel.utils.payload_helper import get_nested_value
from flowork_kernel.ui_shell import shared_properties
from flowork_kernel.ui_shell.components.LabelledCombobox import LabelledCombobox
import ttkbootstrap as ttk
from tkinter import StringVar, filedialog
try:
    import openpyxl
    PANDAS_DEPS_AVAILABLE = True
except ImportError:
    PANDAS_DEPS_AVAILABLE = False
class ExcelManagerModule(BaseModule, IExecutable, IConfigurableUI, IDataPreviewer):
    TIER = "basic"
    def __init__(self, module_id, services):
        super().__init__(module_id, services)
        if not PANDAS_DEPS_AVAILABLE:
            self.logger("FATAL: 'pandas' or 'openpyxl' library is not installed for Excel Manager Module.", "CRITICAL")
    def execute(self, payload: dict, config: dict, status_updater, ui_callback, mode='EXECUTE'):
        if not PANDAS_DEPS_AVAILABLE:
            raise RuntimeError("Required library 'openpyxl' is not installed. Check requirements.txt for this module.")
        if not self.kernel.is_tier_sufficient(self.TIER):
            raise PermissionDeniedError(f"The '{self.manifest.get('name')}' module requires a '{self.TIER}' tier license or higher.")
        operation = config.get('operation', 'read')
        path_mode = config.get('path_mode', 'manual')
        sheet_name = config.get('sheet_name', 0) # Use 0 as default index for pandas
        file_path = ""
        if path_mode == 'dynamic':
            path_key = config.get('path_input_key')
            if not path_key: raise ValueError("Payload key for file path is not set for dynamic mode.")
            file_path = get_nested_value(payload, path_key)
        else: # manual
            file_path = config.get('manual_path')
        if not file_path or (mode == 'EXECUTE' and not os.path.exists(file_path)):
            raise FileNotFoundError(f"Excel file not found at path: {file_path}")
        try:
            if 'data' not in payload or not isinstance(payload['data'], dict):
                payload['data'] = {}
            if operation == 'read':
                status_updater(f"Reading sheet '{sheet_name}' from Excel file...", "INFO")
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                df = df.where(pd.notnull(df), None)
                payload['data']['excel_rows'] = df.to_dict('records')
                status_updater(f"Successfully read {len(df)} rows.", "SUCCESS")
            elif operation == 'write':
                data_key = config.get('data_to_write_key')
                if not data_key: raise ValueError("Payload key for data to write is not configured.")
                data_to_write = get_nested_value(payload, data_key)
                if not isinstance(data_to_write, list):
                    raise TypeError(f"Data to write (from '{data_key}') must be a list of dictionaries.")
                status_updater(f"Writing {len(data_to_write)} new rows to sheet '{sheet_name}'...", "INFO")
                new_df = pd.DataFrame(data_to_write)
                if os.path.exists(file_path):
                    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        try:
                            from openpyxl import load_workbook
                            book = load_workbook(file_path)
                            if sheet_name in book.sheetnames:
                                start_row = book[sheet_name].max_row
                                new_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
                            else: # Sheet doesn't exist, write with header
                                new_df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
                        except Exception:
                            new_df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
                else: # File does not exist, create it
                    with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                        new_df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
                status_updater("Successfully wrote data to Excel file.", "SUCCESS")
            payload['data']['file_path'] = file_path
            return {"payload": payload, "output_name": "success"}
        except Exception as e:
            self.logger(f"Excel module failed: {e}", "ERROR")
            payload['error'] = str(e)
            return {"payload": payload, "output_name": "error"}
    def create_properties_ui(self, parent_frame, get_current_config, available_vars):
        config = get_current_config()
        property_vars = {}
        op_frame = ttk.LabelFrame(parent_frame, text=self.loc.get('prop_operation_label'))
        op_frame.pack(fill='x', padx=5, pady=5)
        op_var = StringVar(value=config.get('operation', 'read'))
        property_vars['operation'] = op_var
        path_source_frame = ttk.LabelFrame(parent_frame, text=self.loc.get('prop_path_source_title'))
        path_source_frame.pack(fill='x', padx=5, pady=5)
        path_mode_var = StringVar(value=config.get('path_mode', 'manual'))
        property_vars['path_mode'] = path_mode_var
        manual_frame = ttk.Frame(path_source_frame)
        dynamic_frame = ttk.Frame(path_source_frame)
        def _toggle_path_mode_ui():
            if path_mode_var.get() == 'manual':
                dynamic_frame.pack_forget(); manual_frame.pack(fill='x', padx=10, pady=5)
            else:
                manual_frame.pack_forget(); dynamic_frame.pack(fill='x', padx=10, pady=5)
        ttk.Radiobutton(path_source_frame, text=self.loc.get('prop_mode_manual'), variable=path_mode_var, value='manual', command=_toggle_path_mode_ui).pack(anchor='w', padx=10)
        ttk.Radiobutton(path_source_frame, text=self.loc.get('prop_mode_dynamic'), variable=path_mode_var, value='dynamic', command=_toggle_path_mode_ui).pack(anchor='w', padx=10)
        manual_path_var = StringVar(value=config.get('manual_path', ''))
        property_vars['manual_path'] = manual_path_var
        ttk.Label(manual_frame, text=self.loc.get('prop_manual_path_label')).pack(anchor='w')
        entry_frame = ttk.Frame(manual_frame)
        entry_frame.pack(fill='x', expand=True)
        ttk.Entry(entry_frame, textvariable=manual_path_var).pack(side='left', fill='x', expand=True)
        ttk.Button(entry_frame, text="...", width=4, command=lambda: manual_path_var.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")]) or manual_path_var.get())).pack(side='left', padx=(5,0))
        path_key_var = StringVar(value=config.get('path_input_key', ''))
        property_vars['path_input_key'] = path_key_var
        LabelledCombobox(dynamic_frame, self.loc.get('prop_path_input_key_label'), path_key_var, list(available_vars.keys()))
        _toggle_path_mode_ui()
        shared_frame = ttk.Frame(parent_frame)
        shared_frame.pack(fill='x', padx=5, pady=5)
        ttk.Label(shared_frame, text=self.loc.get('prop_sheet_name_label')).pack(anchor='w')
        sheet_name_var = StringVar(value=config.get('sheet_name', 'Sheet1'))
        property_vars['sheet_name'] = sheet_name_var
        ttk.Entry(shared_frame, textvariable=sheet_name_var).pack(fill='x')
        write_frame = ttk.Frame(parent_frame) # Initially hidden
        data_to_write_var = StringVar(value=config.get('data_to_write_key', 'data.new_rows'))
        property_vars['data_to_write_key'] = data_to_write_var
        LabelledCombobox(write_frame, self.loc.get('prop_data_to_write_label'), data_to_write_var, list(available_vars.keys()))
        def _toggle_op_ui(*args):
            if op_var.get() == 'write':
                write_frame.pack(fill='x', padx=5, pady=5)
            else:
                write_frame.pack_forget()
        ttk.Radiobutton(op_frame, text=self.loc.get('op_read'), variable=op_var, value='read', command=_toggle_op_ui).pack(side='left', padx=10)
        ttk.Radiobutton(op_frame, text=self.loc.get('op_write'), variable=op_var, value='write', command=_toggle_op_ui).pack(side='left', padx=10)
        _toggle_op_ui()
        ttk.Separator(parent_frame).pack(fill='x', pady=15, padx=5)
        debug_vars = shared_properties.create_debug_and_reliability_ui(parent_frame, config, self.loc)
        property_vars.update(debug_vars)
        return property_vars
    def get_data_preview(self, config: dict):
        if not PANDAS_DEPS_AVAILABLE:
            return [{'error': 'Library `openpyxl` is not installed for this module.'}]
        if config.get('operation', 'read') == 'write':
            return [{'status': 'Will write data from payload to the specified Excel file.'}]
        file_path = config.get('manual_path')
        if not file_path or not os.path.exists(file_path):
            return [{'error': 'Manual path is not set or file does not exist.'}]
        try:
            df = pd.read_excel(file_path, sheet_name=config.get('sheet_name', 0), nrows=5)
            df = df.where(pd.notnull(df), None)
            return df.to_dict('records')
        except Exception as e:
            return [{'error': f"Preview failed: {str(e)}"}]
